#!/usr/bin/env python3
"""
RR-Accounting FastAPI Backend - Entrypoint (moved)

Run: uvicorn backend.app.main:app --reload
"""

#!/usr/bin/env python3
"""
RR-Accounting FastAPI Backend - Reference Implementation
This example shows how to implement REST API endpoints for the accounting system

Run: uvicorn backend.app.main:app --reload
Or: python -m uvicorn backend.app.main:app --reload

Then test with: python api_test_suite.py
"""

from dataclasses import dataclass

from fastapi import FastAPI, HTTPException, Query, Depends, Request, UploadFile, File
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from typing import Optional, List, Any, Callable
from decimal import Decimal
from datetime import datetime
import bcrypt
import csv
import io
from openpyxl import load_workbook

# Import corrected models
from backend.app.models import (
    Base, Company, User, CompanyUser, Customer, Invoice, InvoiceItem,
    Commission, AuditLog, UserRole, InvoiceStatus, CommissionStatus,
    Product, get_database_url
)
from backend.app.permissions import PermissionKey, PERMISSION_MATRIX, RoleName

from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# ============================================================================
# Setup FastAPI & Database
# ============================================================================

app = FastAPI(
    title="RR-Accounting API",
    description="Multi-tenant accounting system with audit trail",
    version="1.0.0"
)

# Add CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins (for development, restrict in production)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create database engine
engine = create_engine(get_database_url(), echo=False)

# Create all tables
Base.metadata.create_all(bind=engine)


@dataclass
class AccessContext:
    user: User
    company_user: CompanyUser
    company_id: int
    role: RoleName

def get_db():
    """Dependency for database sessions."""
    db = Session(engine)
    try:
        yield db
    finally:
        db.close()


def _extract_user_id_from_request(request: Request) -> int:
    auth_header = request.headers.get("Authorization", "")
    raw_token = ""

    if auth_header.lower().startswith("bearer "):
        raw_token = auth_header.split(" ", 1)[1].strip()
    else:
        raw_token = request.query_params.get("token", "")

    if not raw_token.startswith("token_"):
        raise HTTPException(status_code=401, detail="Missing or invalid token")

    user_id_part = raw_token.replace("token_", "", 1)
    if not user_id_part.isdigit():
        raise HTTPException(status_code=401, detail="Invalid token payload")

    return int(user_id_part)


def get_authenticated_user(request: Request, db: Session = Depends(get_db)) -> User:
    user_id = _extract_user_id_from_request(request)
    user = db.query(User).filter(User.id == user_id).first()
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return user


def _resolve_company_id_for_request(
    request: Request,
    user: User,
    db: Session,
) -> int:
    header_company = request.headers.get("X-Company-Id")
    query_company = request.query_params.get("company_id")
    requested_company = header_company or query_company

    if requested_company is not None:
        if not requested_company.isdigit():
            raise HTTPException(status_code=403, detail="Invalid company context")
        return int(requested_company)

    memberships = db.query(CompanyUser).filter(CompanyUser.user_id == user.id).all()
    if len(memberships) == 1:
        return memberships[0].company_id

    raise HTTPException(status_code=403, detail="Company context is required")


def _normalize_role(role: Any) -> RoleName:
    if isinstance(role, UserRole):
        return RoleName(role.value.upper())
    if isinstance(role, str):
        try:
            return RoleName(role.upper())
        except ValueError as exc:
            raise HTTPException(status_code=403, detail="Invalid role") from exc
    raise HTTPException(status_code=403, detail="Invalid role")


def get_access_context(
    request: Request,
    user: User = Depends(get_authenticated_user),
    db: Session = Depends(get_db),
) -> AccessContext:
    company_id = _resolve_company_id_for_request(request, user, db)
    company_user = db.query(CompanyUser).filter(
        CompanyUser.company_id == company_id,
        CompanyUser.user_id == user.id,
    ).first()

    if not company_user:
        raise HTTPException(status_code=403, detail="User has no access to this company")

    db.execute(
        text("SELECT set_config('app.current_company_id', :company_id, true)"),
        {"company_id": str(company_id)},
    )

    role = _normalize_role(company_user.role)
    return AccessContext(user=user, company_user=company_user, company_id=company_id, role=role)


def require_roles(*allowed_roles: Any) -> Callable[[AccessContext], AccessContext]:
    normalized = {
        _normalize_role(role if not isinstance(role, RoleName) else role.value)
        for role in allowed_roles
    }

    def dependency(context: AccessContext = Depends(get_access_context)) -> AccessContext:
        if context.role not in normalized:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return context

    return dependency


def require_owner() -> Callable[[AccessContext], AccessContext]:
    return require_roles(RoleName.OWNER)


def require_permission(permission: PermissionKey) -> Callable[[AccessContext], AccessContext]:
    allowed = PERMISSION_MATRIX[permission]
    return require_roles(*(role.value for role in allowed))

# ============================================================================
# Pydantic Models (Request/Response)
# ============================================================================

class CompanyCreate(BaseModel):
    name: str

class CompanyUpdate(BaseModel):
    name: Optional[str] = None

class CompanyResponse(BaseModel):
    id: int
    name: str
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class CustomerCreate(BaseModel):
    company_id: int
    name: str

class CustomerUpdate(BaseModel):
    name: Optional[str] = None

class CustomerResponse(BaseModel):
    id: int
    company_id: int
    name: str
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class ProductCreate(BaseModel):
    company_id: int
    name: str
    description: Optional[str] = None
    sku: Optional[str] = None
    unit_price: Decimal
    cost_price: Optional[Decimal] = None
    stock_quantity: int = 0
    is_active: bool = True

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    sku: Optional[str] = None
    unit_price: Optional[Decimal] = None
    cost_price: Optional[Decimal] = None
    stock_quantity: Optional[int] = None
    is_active: Optional[bool] = None

class ProductResponse(BaseModel):
    id: int
    company_id: int
    name: str
    description: Optional[str] = None
    sku: Optional[str] = None
    unit_price: Decimal
    cost_price: Optional[Decimal] = None
    stock_quantity: int
    is_active: bool
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class UserCreate(BaseModel):
    email: str
    full_name: str
    password: str

class UserUpdate(BaseModel):
    email: Optional[str] = None
    full_name: Optional[str] = None
    password: Optional[str] = None

class UserResponse(BaseModel):
    id: int
    email: str
    full_name: str
    is_active: bool
    role: Optional[str] = None
    company_id: Optional[int] = None
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class CompanyUserCreate(BaseModel):
    company_id: int
    user_id: int
    commission_percent: Decimal = Field(default="0.00")

class CompanyUserResponse(BaseModel):
    id: int
    company_id: int
    user_id: int
    role: str
    commission_percent: Decimal
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class InvoiceCreate(BaseModel):
    company_id: int
    customer_id: int
    invoice_number: str
    sold_by_user_id: Optional[int] = None
    status: str = "draft"
    total_amount: Decimal = Field(default="0.00")

class InvoiceUpdate(BaseModel):
    status: Optional[str] = None
    total_amount: Optional[Decimal] = None

class InvoiceResponse(BaseModel):
    id: int
    invoice_number: str
    company_id: int
    customer_id: int
    sold_by_user_id: Optional[int]
    created_by_user_id: Optional[int]
    status: str
    total_amount: Decimal
    is_locked: bool
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True

class InvoiceItemCreate(BaseModel):
    invoice_id: int
    description: str
    quantity: Decimal
    unit_price: Decimal
    discount: Decimal = Field(default="0.00")
    total_amount: Decimal

class InvoiceItemResponse(BaseModel):
    id: int
    invoice_id: int
    description: str
    quantity: Decimal
    unit_price: Decimal
    discount: Decimal
    total_amount: Decimal
    
    class Config:
        from_attributes = True

class CommissionResponse(BaseModel):
    id: int
    invoice_id: Optional[int]
    user_id: Optional[int]
    company_id: int
    base_amount: Decimal
    percent: Decimal
    commission_amount: Decimal
    status: str
    created_at: datetime
    updated_at: datetime
    
    class Config:
        from_attributes = True


class CommissionCreate(BaseModel):
    invoice_id: int
    user_id: Optional[int]
    percent: Decimal = Field(default="20.00")
    commission_amount: Decimal


class CommissionUpdate(BaseModel):
    percent: Optional[Decimal] = None
    commission_amount: Optional[Decimal] = None

class AuditLogResponse(BaseModel):
    id: int
    entity_type: str
    entity_id: int
    action: str
    user_id: Optional[int]
    old_data: Optional[dict]
    new_data: Optional[dict]
    created_at: datetime
    
    class Config:
        from_attributes = True

# ============================================================================
# Authentication Models
# ============================================================================

class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    id: int
    email: str
    full_name: str
    is_active: bool
    role: Optional[str] = None
    company_id: Optional[int] = None
    token: str = "token"  # In real app, this would be JWT
    
    class Config:
        from_attributes = True

class MeResponse(BaseModel):
    id: int
    email: str
    full_name: str
    is_active: bool
    role: Optional[str] = None
    company_id: Optional[int] = None
    
    class Config:
        from_attributes = True

# ============================================================================
# Health Check
# ============================================================================

@app.get("/health", status_code=200, response_model=None)
async def health_check(db: Session = Depends(get_db)):
    """Health check endpoint with database connectivity test."""
    try:
        # Test database connection
        db.execute(text("SELECT 1"))
        return {
            "status": "ok",
            "database": "connected",
            "message": "RR-Accounting API is running",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "status": "error",
                "database": "disconnected",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }
        )

# ============================================================================
# Authentication Endpoints
# ============================================================================

@app.post("/api/auth/login", response_model=LoginResponse, status_code=200)
async def login(credentials: LoginRequest, db: Any = None):
    """Login user with email and password."""
    if db is None:
        db = next(get_db())
    
    # Find user by email
    user = db.query(User).filter(User.email == credentials.email).first()
    
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user.is_active:
        raise HTTPException(status_code=401, detail="User account is inactive")
    
    # Verify password
    try:
        if not bcrypt.checkpw(credentials.password.encode('utf-8'), user.password_hash.encode('utf-8')):
            raise HTTPException(status_code=401, detail="Invalid email or password")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    company_user = db.query(CompanyUser).filter(CompanyUser.user_id == user.id).first()

    # Return user info with token
    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "is_active": user.is_active,
        "role": company_user.role.value.upper() if company_user else None,
        "company_id": company_user.company_id if company_user else None,
        "token": f"token_{user.id}"  # Simple token format
    }

@app.get("/api/auth/me", response_model=MeResponse)
async def get_current_user(user_id: int = Query(...), token: str = Query(...), db: Any = None):
    """Get current logged-in user info."""
    if db is None:
        db = next(get_db())
    
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    if not user.is_active:
        raise HTTPException(status_code=401, detail="User account is inactive")
    
    company_user = db.query(CompanyUser).filter(CompanyUser.user_id == user.id).first()

    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "is_active": user.is_active,
        "role": company_user.role.value.upper() if company_user else None,
        "company_id": company_user.company_id if company_user else None,
    }

# ============================================================================
# Company Endpoints
# ============================================================================

@app.post("/api/companies", response_model=CompanyResponse, status_code=201)
async def create_company(
    company: CompanyCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Create a new company."""
    
    new_company = Company(name=company.name)
    db.add(new_company)
    db.flush()

    owner_membership = CompanyUser(
        company_id=new_company.id,
        user_id=context.user.id,
        role=UserRole.OWNER,
        commission_percent=Decimal("0.00"),
    )
    db.add(owner_membership)

    db.commit()
    db.refresh(new_company)
    return new_company

@app.get("/api/companies", response_model=List[CompanyResponse])
async def list_companies(
    skip: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(get_access_context),
):
    """List all companies."""

    companies = (
        db.query(Company)
        .join(CompanyUser, CompanyUser.company_id == Company.id)
        .filter(CompanyUser.user_id == context.user.id)
        .offset(skip)
        .limit(limit)
        .all()
    )
    return companies

@app.get("/api/companies/{company_id}", response_model=CompanyResponse)
async def get_company(
    company_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(get_access_context),
):
    """Get a company by ID."""
    if company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return company

@app.put("/api/companies/{company_id}", response_model=CompanyResponse)
async def update_company(
    company_id: int,
    company_update: CompanyUpdate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Update a company."""
    if company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    if company_update.name:
        company.name = company_update.name
    
    db.commit()
    db.refresh(company)
    return company


@app.delete("/api/companies/{company_id}", status_code=204)
async def delete_company(
    company_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    if company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")

    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    db.delete(company)
    db.commit()
    return None

# ============================================================================
# Customer Endpoints
# ============================================================================

@app.post("/api/customers", response_model=CustomerResponse, status_code=201)
async def create_customer(
    customer: CustomerCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.CUSTOMER_CREATE)),
):
    """Create a new customer."""
    if customer.company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    new_customer = Customer(
        company_id=customer.company_id,
        name=customer.name
    )
    db.add(new_customer)
    db.commit()
    db.refresh(new_customer)
    return new_customer

@app.get("/api/customers", response_model=List[CustomerResponse])
async def list_customers(
    company_id: Optional[int] = None,
    skip: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.CUSTOMER_READ)),
):
    """List customers."""
    if company_id is not None and company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    query = db.query(Customer).filter(Customer.company_id == context.company_id)
    
    customers = query.offset(skip).limit(limit).all()
    return customers

@app.get("/api/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.CUSTOMER_READ)),
):
    """Get a customer by ID."""
    
    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.company_id == context.company_id,
    ).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@app.put("/api/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(
    customer_id: int,
    customer_update: CustomerUpdate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.CUSTOMER_UPDATE)),
):
    """Update a customer."""
    
    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.company_id == context.company_id,
    ).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    if customer_update.name:
        customer.name = customer_update.name
    
    db.commit()
    db.refresh(customer)
    return customer


@app.delete("/api/customers/{customer_id}", status_code=204)
async def delete_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.CUSTOMER_DELETE)),
):
    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.company_id == context.company_id,
    ).first()
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    db.delete(customer)
    db.commit()
    return None

# ============================================================================
# Product Endpoints
# ============================================================================

@app.post("/api/products", response_model=ProductResponse, status_code=201)
async def create_product(
    product: ProductCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_CREATE)),
):
    """Create a new product."""
    if product.company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    new_product = Product(
        company_id=product.company_id,
        name=product.name,
        description=product.description,
        sku=product.sku,
        unit_price=product.unit_price,
        cost_price=product.cost_price,
        stock_quantity=product.stock_quantity,
        is_active=product.is_active,
    )
    db.add(new_product)
    db.commit()
    db.refresh(new_product)
    return new_product

@app.get("/api/products", response_model=List[ProductResponse])
async def list_products(
    company_id: Optional[int] = None,
    skip: int = Query(0),
    limit: int = Query(100),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_READ)),
):
    """List products."""
    if company_id is not None and company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    query = db.query(Product).filter(Product.company_id == context.company_id)
    
    products = query.offset(skip).limit(limit).all()
    return products

# ============================================================================
# Product Search Endpoints (for Invoice autocomplete)
# ============================================================================

class ProductSearchResponse(BaseModel):
    """Minimal product response for search/autocomplete."""
    id: int
    name: str
    sku: Optional[str]
    unit_price: float

    class Config:
        from_attributes = True

@app.get("/api/products/search-suggestions", response_model=List[ProductSearchResponse])
async def search_products(
    q: str = "",
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_READ)),
):
    """
    Search products by name or SKU (partial, case-insensitive).
    Returns max 10 results for performance.
    
    Query Parameters:
    - q: Search string (min 1 char, searches both name and sku)
    
    Example: GET /api/products/search-suggestions?q=laptop
    Returns: List[ProductSearchResponse]
    """
    if not q or len(q.strip()) < 1:
        return []
    
    search_pattern = f"%{q.strip().lower()}%"
    
    products = db.query(Product).filter(
        Product.company_id == context.company_id,
        Product.is_active == True,
    ).filter(
        (Product.name.ilike(search_pattern)) |
        (Product.sku.ilike(search_pattern))
    ).limit(10).all()
    
    return [
        {
            "id": product.id,
            "name": product.name,
            "sku": product.sku,
            "unit_price": float(product.unit_price),
        }
        for product in products
    ]

@app.get("/api/products/by-code/{code}", response_model=ProductSearchResponse)
async def get_product_by_code(
    code: str,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_READ)),
):
    """
    Get product by SKU (product code) - exact match.
    Used when user enters a code and presses Enter or blurs the field.
    
    Returns: ProductSearchResponse with id, name, sku, unit_price
    Raises: 404 if not found
    """
    product = db.query(Product).filter(
        Product.sku == code.strip(),
        Product.company_id == context.company_id,
        Product.is_active == True,
    ).first()
    
    if not product:
        raise HTTPException(status_code=404, detail=f"Product with code '{code}' not found")
    
    return {
        "id": product.id,
        "name": product.name,
        "sku": product.sku,
        "unit_price": float(product.unit_price),
    }

@app.post("/api/products/import", status_code=201)
async def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_IMPORT)),
):
    """Import products from CSV or XLSX file.
    
    Expected columns:
    - name (required)
    - description (optional)
    - sku (optional)
    - unit_price (required)
    - cost_price (optional)
    - stock_quantity (optional, default: 0)
    - is_active (optional, default: true)
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    filename = file.filename.lower()
    
    try:
        contents = await file.read()
        
        products_data = []
        
        if filename.endswith('.csv'):
            # Parse CSV
            text_content = contents.decode('utf-8-sig')  # Handle BOM
            csv_reader = csv.DictReader(io.StringIO(text_content))
            products_data = list(csv_reader)
            
        elif filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            wb = load_workbook(filename=io.BytesIO(contents), read_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip empty rows
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    products_data.append(row_dict)
        else:
            raise HTTPException(
                status_code=400, 
                detail="Invalid file format. Only CSV, XLS, and XLSX are supported"
            )
        
        # Validate and create products
        created_products = []
        errors = []
        
        for idx, row in enumerate(products_data, start=2):  # Start at 2 for spreadsheet row numbers
            try:
                # Validate required fields
                if not row.get('name'):
                    errors.append(f"Row {idx}: 'name' is required")
                    continue
                
                if not row.get('unit_price'):
                    errors.append(f"Row {idx}: 'unit_price' is required")
                    continue
                
                # Parse values
                unit_price = Decimal(str(row['unit_price']))
                cost_price = Decimal(str(row['cost_price'])) if row.get('cost_price') else None
                stock_quantity = int(row.get('stock_quantity', 0))
                
                # Parse is_active (default to True)
                is_active_str = str(row.get('is_active', 'true')).lower()
                is_active = is_active_str in ('true', '1', 'yes', 't', 'y')
                
                # Create product
                product = Product(
                    company_id=context.company_id,
                    name=row['name'],
                    description=row.get('description'),
                    sku=row.get('sku'),
                    unit_price=unit_price,
                    cost_price=cost_price,
                    stock_quantity=stock_quantity,
                    is_active=is_active,
                )
                
                db.add(product)
                created_products.append(row['name'])
                
            except ValueError as e:
                errors.append(f"Row {idx}: Invalid number format - {str(e)}")
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        if created_products:
            db.commit()
        
        return {
            "success": True,
            "imported": len(created_products),
            "errors": errors,
            "message": f"Successfully imported {len(created_products)} products"
        }
        
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="File encoding error. Please use UTF-8 encoding")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# ============================================================================
# Product ID-based Endpoints (dynamic routes - defined last to avoid shadowing)
# ============================================================================

@app.get("/api/products/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_READ)),
):
    """Get a product by ID."""
    
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.company_id == context.company_id,
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product

@app.put("/api/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: int,
    product_update: ProductUpdate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_UPDATE)),
):
    """Update a product."""
    
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.company_id == context.company_id,
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    if product_update.name is not None:
        product.name = product_update.name
    if product_update.description is not None:
        product.description = product_update.description
    if product_update.sku is not None:
        product.sku = product_update.sku
    if product_update.unit_price is not None:
        product.unit_price = product_update.unit_price
    if product_update.cost_price is not None:
        product.cost_price = product_update.cost_price
    if product_update.stock_quantity is not None:
        product.stock_quantity = product_update.stock_quantity
    if product_update.is_active is not None:
        product.is_active = product_update.is_active
    
    db.commit()
    db.refresh(product)
    return product

@app.delete("/api/products/{product_id}", status_code=204)
async def delete_product(
    product_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.PRODUCT_DELETE)),
):
    """Delete a product."""
    product = db.query(Product).filter(
        Product.id == product_id,
        Product.company_id == context.company_id,
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    db.delete(product)
    db.commit()
    return None
# ============================================================================
# User Endpoints
# ============================================================================

@app.post("/api/users", response_model=UserResponse, status_code=201)
async def create_user(
    user: UserCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Create a new user."""
    
    # Hash the password
    password_hash = bcrypt.hashpw(user.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    new_user = User(
        email=user.email,
        full_name=user.full_name,
        password_hash=password_hash,
        is_active=True
    )
    db.add(new_user)
    db.flush()

    membership = CompanyUser(
        company_id=context.company_id,
        user_id=new_user.id,
        role=UserRole.ACCOUNTANT,
    )
    db.add(membership)
    db.commit()
    db.refresh(new_user)
    db.refresh(membership)
    return {
        "id": new_user.id,
        "email": new_user.email,
        "full_name": new_user.full_name,
        "is_active": new_user.is_active,
        "role": membership.role.value.upper(),
        "company_id": membership.company_id,
        "created_at": new_user.created_at,
        "updated_at": new_user.updated_at,
    }

@app.get("/api/users", response_model=List[UserResponse])
async def list_users(
    skip: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """List all users."""
    pairs = db.query(User, CompanyUser).join(
        CompanyUser,
        CompanyUser.user_id == User.id,
    ).filter(
        CompanyUser.company_id == context.company_id,
    ).offset(skip).limit(limit).all()

    return [
        {
            "id": user.id,
            "email": user.email,
            "full_name": user.full_name,
            "is_active": user.is_active,
            "role": company_user.role.value.upper(),
            "company_id": company_user.company_id,
            "created_at": user.created_at,
            "updated_at": user.updated_at,
        }
        for user, company_user in pairs
    ]

@app.get("/api/users/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Get a user by ID."""
    pair = db.query(User, CompanyUser).join(
        CompanyUser,
        CompanyUser.user_id == User.id,
    ).filter(
        User.id == user_id,
        CompanyUser.company_id == context.company_id,
    ).first()
    if not pair:
        raise HTTPException(status_code=404, detail="User not found")
    user, company_user = pair
    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "is_active": user.is_active,
        "role": company_user.role.value.upper(),
        "company_id": company_user.company_id,
        "created_at": user.created_at,
        "updated_at": user.updated_at,
    }

@app.put("/api/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user_update: UserUpdate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Update a user with optional password rotation."""
    pair = db.query(User, CompanyUser).join(
        CompanyUser,
        CompanyUser.user_id == User.id,
    ).filter(
        User.id == user_id,
        CompanyUser.company_id == context.company_id,
    ).first()
    if not pair:
        raise HTTPException(status_code=404, detail="User not found")
    user, company_user = pair

    if user_update.email is not None:
        user.email = user_update.email
    if user_update.full_name is not None:
        user.full_name = user_update.full_name
    if user_update.password:
        user.password_hash = bcrypt.hashpw(
            user_update.password.encode('utf-8'),
            bcrypt.gensalt()
        ).decode('utf-8')

    db.commit()
    db.refresh(user)
    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "is_active": user.is_active,
        "role": company_user.role.value.upper(),
        "company_id": company_user.company_id,
        "created_at": user.created_at,
        "updated_at": user.updated_at,
    }

@app.delete("/api/users/{user_id}", status_code=204)
async def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Delete a user (preserves invoices, commissions via FK SET NULL)."""
    user = db.query(User).join(
        CompanyUser,
        CompanyUser.user_id == User.id,
    ).filter(
        User.id == user_id,
        CompanyUser.company_id == context.company_id,
    ).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.delete(user)
    db.commit()
    return None

# ============================================================================
# Company User Endpoints
# ============================================================================

@app.post("/api/company-users", response_model=CompanyUserResponse, status_code=201)
async def assign_user_to_company(
    company_user: CompanyUserCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    """Assign user to company with commission percentage."""
    if company_user.company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    new_company_user = CompanyUser(
        company_id=company_user.company_id,
        user_id=company_user.user_id,
        commission_percent=company_user.commission_percent
    )
    db.add(new_company_user)
    db.commit()
    db.refresh(new_company_user)
    return new_company_user

# ============================================================================
# Invoice Endpoints
# ============================================================================

@app.post("/api/invoices", response_model=InvoiceResponse, status_code=201)
async def create_invoice(
    invoice: InvoiceCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.INVOICE_CREATE)),
):
    """Create a new invoice."""
    if invoice.company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")
    
    status_map = {
        "draft": "draft",
        "sent": "issued",
        "issued": "issued",
        "overdue": "issued",
        "paid": "paid",
    }

    new_invoice = Invoice(
        company_id=invoice.company_id,
        customer_id=invoice.customer_id,
        invoice_number=invoice.invoice_number,
        sold_by_user_id=invoice.sold_by_user_id,
        created_by_user_id=context.user.id,
        status=status_map.get(invoice.status, "draft"),
        total_amount=invoice.total_amount,
    )
    db.add(new_invoice)
    db.commit()
    db.refresh(new_invoice)
    return new_invoice

@app.get("/api/invoices", response_model=List[InvoiceResponse])
async def list_invoices(
    company_id: Optional[int] = None,
    sold_by_user_id: Optional[int] = None,
    skip: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.INVOICE_READ)),
):
    """List invoices."""
    if company_id is not None and company_id != context.company_id:
        raise HTTPException(status_code=403, detail="Cross-company access denied")

    query = db.query(Invoice).filter(Invoice.company_id == context.company_id)
    if context.role == RoleName.SALES:
        query = query.filter(Invoice.sold_by_user_id == context.user.id)
    elif sold_by_user_id:
        query = query.filter(Invoice.sold_by_user_id == sold_by_user_id)
    
    invoices = query.offset(skip).limit(limit).all()
    return invoices

@app.get("/api/invoices/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.INVOICE_READ)),
):
    """Get an invoice by ID."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if context.role == RoleName.SALES and invoice.sold_by_user_id != context.user.id:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    return invoice

@app.put("/api/invoices/{invoice_id}", response_model=InvoiceResponse)
async def update_invoice(
    invoice_id: int,
    invoice_update: InvoiceUpdate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.INVOICE_UPDATE)),
):
    """Update an invoice (triggers commission snapshot if status → PAID)."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if invoice.is_locked and context.role != RoleName.OWNER:
        raise HTTPException(status_code=403, detail="Only owner can edit locked invoice")
    
    if invoice_update.status:
        status_map = {
            "draft": "draft",
            "sent": "issued",
            "issued": "issued",
            "overdue": "issued",
            "paid": "paid",
        }
        invoice.status = status_map.get(invoice_update.status, invoice.status)
    if invoice_update.total_amount is not None:
        invoice.total_amount = invoice_update.total_amount
    
    db.commit()
    db.refresh(invoice)
    return invoice


@app.delete("/api/invoices/{invoice_id}", status_code=204)
async def delete_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.INVOICE_DELETE)),
):
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    status_value = invoice.status.value if isinstance(invoice.status, InvoiceStatus) else str(invoice.status).lower()
    if context.role == RoleName.ACCOUNTANT and status_value == InvoiceStatus.PAID.value:
        raise HTTPException(status_code=403, detail="Accountant cannot delete paid invoices")

    if invoice.is_locked and context.role != RoleName.OWNER:
        raise HTTPException(status_code=403, detail="Only owner can delete locked invoice")

    db.delete(invoice)
    db.commit()
    return None


@app.post("/api/invoices/{invoice_id}/lock", response_model=InvoiceResponse)
async def lock_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    invoice.is_locked = True
    db.commit()
    db.refresh(invoice)
    return invoice


@app.post("/api/invoices/{invoice_id}/unlock", response_model=InvoiceResponse)
async def unlock_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    invoice.is_locked = False
    db.commit()
    db.refresh(invoice)
    return invoice

# ============================================================================
# Invoice Item Endpoints
# ============================================================================

@app.post("/api/invoice-items", response_model=InvoiceItemResponse, status_code=201)
async def create_invoice_item(
    item: InvoiceItemCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_roles(RoleName.OWNER.value, RoleName.ACCOUNTANT.value)),
):
    """Create an invoice item with validation (Issue #1)."""
    invoice = db.query(Invoice).filter(
        Invoice.id == item.invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if invoice.is_locked and context.role != RoleName.OWNER:
        raise HTTPException(status_code=403, detail="Only owner can modify locked invoice")
    
    # Validate total_amount = quantity * unit_price - discount
    expected_total = (item.quantity * item.unit_price) - item.discount
    expected_total = expected_total.quantize(Decimal('0.01'))
    
    if item.total_amount != expected_total:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid total_amount. Expected {expected_total}, got {item.total_amount}"
        )
    
    new_item = InvoiceItem(
        invoice_id=item.invoice_id,
        description=item.description,
        quantity=item.quantity,
        unit_price=item.unit_price,
        discount=item.discount,
        total_amount=item.total_amount
    )
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    return new_item

@app.get("/api/invoice-items")
async def list_invoice_items(
    invoice_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.INVOICE_READ)),
):
    """List invoice items."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if context.role == RoleName.SALES and invoice.sold_by_user_id != context.user.id:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice_id).all()
    return items

# ============================================================================
# Commission Endpoints
# ============================================================================

@app.get("/api/commissions", response_model=List[CommissionResponse])
async def list_commissions(
    invoice_id: Optional[int] = None,
    user_id: Optional[int] = None,
    skip: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.COMMISSION_READ)),
):
    """List commissions."""
    query = db.query(Commission).filter(Commission.company_id == context.company_id)
    if invoice_id:
        query = query.filter(Commission.invoice_id == invoice_id)
    if context.role == RoleName.SALES:
        query = query.filter(Commission.user_id == context.user.id)
    elif user_id:
        query = query.filter(Commission.user_id == user_id)
    
    commissions = query.offset(skip).limit(limit).all()
    return commissions

@app.get("/api/commissions/{commission_id}", response_model=CommissionResponse)
async def get_commission(
    commission_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.COMMISSION_READ)),
):
    """Get a commission by ID."""
    commission = db.query(Commission).filter(
        Commission.id == commission_id,
        Commission.company_id == context.company_id,
    ).first()
    if not commission:
        raise HTTPException(status_code=404, detail="Commission not found")

    if context.role == RoleName.SALES and commission.user_id != context.user.id:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    return commission


@app.post("/api/commissions", response_model=CommissionResponse, status_code=201)
async def create_commission(
    payload: CommissionCreate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    invoice = db.query(Invoice).filter(
        Invoice.id == payload.invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    new_commission = Commission(
        company_id=context.company_id,
        invoice_id=payload.invoice_id,
        user_id=payload.user_id,
        base_amount=Decimal(str(invoice.total_amount)),
        percent=payload.percent,
        commission_amount=payload.commission_amount,
        status=CommissionStatus.PENDING,
    )
    db.add(new_commission)
    db.commit()
    db.refresh(new_commission)
    return new_commission


@app.put("/api/commissions/{commission_id}", response_model=CommissionResponse)
async def update_commission(
    commission_id: int,
    payload: CommissionUpdate,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    commission = db.query(Commission).filter(
        Commission.id == commission_id,
        Commission.company_id == context.company_id,
    ).first()
    if not commission:
        raise HTTPException(status_code=404, detail="Commission not found")

    if payload.percent is not None:
        commission.percent = payload.percent
    if payload.commission_amount is not None:
        commission.commission_amount = payload.commission_amount

    db.commit()
    db.refresh(commission)
    return commission


@app.delete("/api/commissions/{commission_id}", status_code=204)
async def delete_commission(
    commission_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    commission = db.query(Commission).filter(
        Commission.id == commission_id,
        Commission.company_id == context.company_id,
    ).first()
    if not commission:
        raise HTTPException(status_code=404, detail="Commission not found")

    db.delete(commission)
    db.commit()
    return None

@app.post("/api/commissions/{commission_id}/approve", response_model=CommissionResponse)
async def approve_commission(
    commission_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    commission = db.query(Commission).filter(
        Commission.id == commission_id,
        Commission.company_id == context.company_id,
    ).first()
    if not commission:
        raise HTTPException(status_code=404, detail="Commission not found")

    commission.status = CommissionStatus.APPROVED
    db.commit()
    db.refresh(commission)
    return commission


@app.post("/api/commissions/{commission_id}/mark-paid", response_model=CommissionResponse)
async def mark_commission_paid(
    commission_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_owner()),
):
    commission = db.query(Commission).filter(
        Commission.id == commission_id,
        Commission.company_id == context.company_id,
    ).first()
    if not commission:
        raise HTTPException(status_code=404, detail="Commission not found")

    commission.status = CommissionStatus.PAID
    db.commit()
    db.refresh(commission)
    return commission


@app.post(
    "/api/invoices/{invoice_id}/create-commission-snapshot",
    response_model=List[CommissionResponse],
    status_code=201,
)
async def create_commission_snapshot(
    invoice_id: int,
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.COMMISSION_CREATE_SNAPSHOT)),
):
    """Create commission snapshot(s) for an invoice if not already created."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.company_id == context.company_id,
    ).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if context.role == RoleName.SALES:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    existing = db.query(Commission).filter(
        Commission.invoice_id == invoice_id,
        Commission.company_id == context.company_id,
    ).all()
    if existing:
        return existing

    default_percent = Decimal("20.00")
    percent = default_percent

    if invoice.sold_by_user_id:
        company_user = db.query(CompanyUser).filter(
            CompanyUser.company_id == invoice.company_id,
            CompanyUser.user_id == invoice.sold_by_user_id,
        ).first()
        if company_user and company_user.commission_percent is not None:
            percent = Decimal(str(company_user.commission_percent))

    base_amount = Decimal(str(invoice.total_amount))
    commission_amount = (base_amount * percent / Decimal("100")).quantize(
        Decimal("0.01")
    )

    snapshot = Commission(
        company_id=invoice.company_id,
        invoice_id=invoice.id,
        user_id=invoice.sold_by_user_id,
        base_amount=base_amount,
        percent=percent,
        commission_amount=commission_amount,
        status=CommissionStatus.PENDING,
    )
    db.add(snapshot)
    db.commit()

    return db.query(Commission).filter(
        Commission.invoice_id == invoice_id,
        Commission.company_id == context.company_id,
    ).all()

# ============================================================================
# Audit Log Endpoints
# ============================================================================

@app.get("/api/audit-logs", response_model=List[AuditLogResponse])
async def list_audit_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    skip: int = Query(0),
    limit: int = Query(10),
    db: Session = Depends(get_db),
    context: AccessContext = Depends(require_permission(PermissionKey.AUDIT_READ)),
):
    """List audit logs."""
    query = db.query(AuditLog).filter(AuditLog.company_id == context.company_id)
    if entity_type:
        query = query.filter(AuditLog.entity_type == entity_type)
    if entity_id:
        query = query.filter(AuditLog.entity_id == entity_id)
    
    logs = query.order_by(AuditLog.created_at.desc()).offset(skip).limit(limit).all()
    return logs

# ============================================================================
# Main Entry Point
# ============================================================================

if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host='127.0.0.1', port=8000)
